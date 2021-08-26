import mgrs
import openpyxl

file_name = '미추홀구 방범용 CCTV 통합 리스트 03.16.xlsx'
# load excel file
wb = openpyxl.load_workbook(file_name)

ws = wb.active
ws.insert_cols(7, 1)

for x in range(4, 50):
    if ws.cell(x, 6).value != None:
        # get the value of the lat and long value
        strings = ws.cell(x, 6).value
        cut_strings = strings.split(',')
        print(cut_strings)

        # get latitude and longtitude

        lat = cut_strings[0]
        long = cut_strings[1]

        m = mgrs.MGRS()

        c = m.toMGRS(lat, long)
        print(c)

        ws.cell(x, 7).value = c


wb.save(file_name)

wb.close()
