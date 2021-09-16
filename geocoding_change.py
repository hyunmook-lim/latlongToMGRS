import googlemaps
from datetime import datetime
import pprint
import mgrs
import openpyxl


# get googlemaps geocoding API key
gmaps = googlemaps.Client(key='AIzaSyB8OVEKyCF0Ntu6Mew-piV-P_-zmTzexe0')


def change_MGRS(file_name):
    print('this funcion is started')
    print(file_name)

    file_name = str(file_name).replace('/', '\\')

    print(file_name)

    # get file
    # load excel file
    wb = openpyxl.load_workbook(file_name)

    ws = wb.active
    # for end of the address' cells, get the lat and lng

    for x in range(2, ws.max_row + 1):

        print(ws.cell(x, 1).value)

        address = ws.cell(x, 1).value

        geocode_result = gmaps.geocode(address)

        # if the json has no value, just get next cell
        if geocode_result != []:
            # put the lat and lng at the excel file
            ws.cell(
                x, 2).value = geocode_result[0]['geometry']['location']['lng']
            ws.cell(
                x, 3).value = geocode_result[0]['geometry']['location']['lat']

            # change lat and lng to MGRS and put the value at the MGRS columns
            lat = ws.cell(x, 3).value
            lng = ws.cell(x, 2).value

        m = mgrs.MGRS()

        c = m.toMGRS(lat, lng)
        print(c)

        ws.cell(x, 4).value = c

    wb.save(file_name)
    wb.close()
